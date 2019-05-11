import os
import pandas as pd
import calendar
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# Let user choose which sheet to record expenses
# https://stackoverflow.com/questions/34927479/command-line-show-list-of-options-and-let-user-choose
def choose_sheet(sheetnames):
    print("Please choose which month to record the expenses :")
    for idx, element in enumerate(sheetnames):
        print("{}) {}".format(idx+1, element))
    i = input("Enter number: ")
    try:
        if 0 < int(i) <= len(sheetnames):
            return int(i)
    except:
        pass
    return None


# Let user choose which categories the expenses belongs to
def choose_category(categories):
    print("Please choose the appropriate category for the expenses :")
    for idx, element in enumerate(categories):
        print("{}) {}".format(idx+1, element))
    i = input("Enter number: ")
    try:
        if 0 < int(i) <= len(categories):
            return int(i)
    except:
        pass
    return None


if os.path.isfile("Expenses.xlsx"):
    workbook = load_workbook(filename="Expenses.xlsx")
    sheet_name = choose_sheet(workbook.sheetnames)
    print(workbook.sheetnames[sheet_name - 1] + " is selected.")
    sheet_name_df = pd.read_excel("Expenses.xlsx", sheet_name=workbook.sheetnames[sheet_name - 1])
    items = int(input("How many items would you like to record? :"))

    # Expenses categories
    # https://localfirstbank.com/content/personal-budget-categories/
    categories = ['Housing', 'Transportation', 'Food', 'Utilities', 'Clothing', 'Medical/ Healthcare',
                  'Insurance', 'Household Items/ Supplies', 'Personal', 'Debt', 'Retirement',
                  'Education', 'Savings', 'Gifts/ Donations', 'Entertainment']

    for i in range(items):
        Date = input("What is the date of purchase? (YYYY-MM-DD) :")
        Item = input("What did you buy? :")
        Quantity = int(input("How many did you buy? :"))
        Amount = float(input("How much is it? :"))
        Category = choose_category(categories)
        print(categories[Category - 1] + " is selected.")
        chosen_category = categories[Category - 1]

        sheet_name_df = sheet_name_df.append({'Date': Date, 'Item': Item, 'Quantity': Quantity, 'Amount': Amount,
                                              'Category': chosen_category}, ignore_index=True)
        i += 1
        print("Here is your updated expenses record.")
        print(sheet_name_df)

    write = pd.ExcelWriter('Expenses.xlsx', engine="openpyxl")
    write.book = workbook
    # https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    write.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    sheet_name_df.to_excel(write, workbook.sheetnames[sheet_name - 1], index=None)
    write.save()
    workbook.close()

    print("The record is saved successfully.")
    if input("Would you like to see how much you have spent in each category? (Y/N) :") == "Y":
        df = pd.read_excel("Expenses.xlsx", sheet_name=workbook.sheetnames[sheet_name - 1])
        amount = df.Amount
        category = df.Category.unique()

        # Save the sum of amount of a category after filtering
        sum_ls = []

        if len(category) > 1:
            for i in range(0, len(category)):
                sum_ls.append(sum(df[df.Category == category[i]].Amount))

        # Create the donut chart
        explode = [0.015]*len(category)

        fig1, ax1 = plt.subplots()
        ax1.pie(sum_ls, explode=explode, labels=category, autopct='%1.1f%%',
                shadow=True, startangle=90)
        ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        ax1.title.set_text("Expenses in " + workbook.sheetnames[sheet_name - 1])
        my_circle = plt.Circle((0, 0), 0.8, color='white')
        fig1.gca().add_artist(my_circle)
        plt.show()
    else:
        print("Have a nice day.")

else:
    print("'Expenses.xlsx' is not found!")
    if input("Would you like me to create one for you? (Y/N) :") == "Y":
        expenses = openpyxl.Workbook()
        for i in range(1, 13):
            month_name = expenses.create_sheet(str(calendar.month_name[i]))
        expenses.remove_sheet(expenses.get_sheet_by_name("Sheet"))  # Remove the first sheet
        expenses.save("Expenses.xlsx")

        # Create column name in each sheet
        month = pd.DataFrame(columns=["Date", "Item", "Quantity", "Amount", "Category"])
        write = pd.ExcelWriter("Expenses.xlsx")
        for i in range(1, 13):
            month.to_excel(write, str(calendar.month_name[i]), index=None)
        write.save()
        expenses.close()
        print("Expenses Tracker is created successfully.")
    else:
        print("Have a nice day.")
